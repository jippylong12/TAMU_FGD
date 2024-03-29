_author_ = "Marcus Salinas"

from openpyxl import Workbook
from collections import defaultdict


def outputData(masterDictionary, title):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = "Course"
    ws.cell(row=1, column=2).value = "Professor"
    ws.cell(row=1, column=3).value = "GPA"
    ws.cell(row=1, column=4).value = "Num of A's"
    ws.cell(row=1, column=5).value = "Num of B's"
    ws.cell(row=1, column=6).value = "Num of C's"
    ws.cell(row=1, column=7).value = "Num of D's"
    ws.cell(row=1, column=8).value = "Num of F's"
    ws.cell(row=1, column=9).value = "Num of Q Drop's"

    rowCount = 2

    #'{:.1%}'.format(1/3.0)

    # added a sort
    sortThisList = []

    # add all the keys to the list
    for key in masterDictionary.keys():
        sortThisList.append(key)

    # sort the list
    sortThisList.sort()

    # writing class data to spreadsheet
    # for each class in the list
    for course in sortThisList:
        dataList = masterDictionary[course]  # find it in the masterDictionary
        # output the data
        ws.cell(row=rowCount, column=1).value = course
        rowCount += 1
        for item in dataList:
            professorName = list(item.keys())[0]
            ws.cell(row=rowCount, column=2).value = professorName
            ws.cell(row=rowCount, column=3).value = item[professorName][15]
            ws.cell(row=rowCount, column=4).value = item[professorName][0]
            ws.cell(row=rowCount, column=5).value = item[professorName][1]
            ws.cell(row=rowCount, column=6).value = item[professorName][2]
            ws.cell(row=rowCount, column=7).value = item[professorName][3]
            ws.cell(row=rowCount, column=8).value = item[professorName][4]
            ws.cell(row=rowCount, column=9).value = item[professorName][7] # Q Drop
            rowCount += 1
        rowCount += 1

    return wb
